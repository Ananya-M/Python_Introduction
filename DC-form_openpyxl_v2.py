import tkinter as tk
import xlsxwriter

from openpyxl import Workbook
from functools import partial

root=tk.Tk()
root.title("DC")
sln=[]
hsn=[]
desc=[]
qty=[]
uom=[]
rate=[]
con=[]

def add(e1):
    n=e1.get()
    x=int(n)
    print(x)
    i=1
    #master=tk.Toplevel(root)
    tk.Label(root,text="Payment/Delivery Terms:").grid(row=2,column=0)
    e7=tk.Entry(root)
    e7.grid(row=2,column=1)
    tk.Label(root,text="Consignee:").grid(row=1,column=3)
    p=1
    while (p<=7):
        ent=tk.Entry(root)
        ent.grid(row=p+1,column=3)
        con.append(ent)
        p+=1
    
    tk.Label(root,text="S.N.").grid(row=9,column=0)
    tk.Label(root,text="HSN Code").grid(row=9,column=1)
    tk.Label(root,text="Goods/ Services Description").grid(row=9,column=2)
    tk.Label(root,text="Qty").grid(row=9,column=3)
    tk.Label(root,text="UOM").grid(row=9,column=4)
    tk.Label(root,text="Unit Rate (INR)").grid(row=9,column=5)
    
    while(i<=x):
            sln.append(i)
            tk.Label(root,text=i).grid(row=i+9,column=0)
            e2=tk.Entry(root)
            e2.grid(row=i+9,column=1)
            hsn.append(e2)
            e3=tk.Entry(root)
            e3.grid(row=i+9,column=2)
            desc.append(e3)
            e4=tk.Entry(root)
            e4.grid(row=i+9,column=3)
            qty.append(e4)
            e5=tk.Entry(root)
            e5.grid(row=i+9,column=4)
            uom.append(e5)
            e6=tk.Entry(root)
            e6.grid(row=i+9,column=5)
            rate.append(e6)
            i+=1
    tk.Button(root,text='Next',command=partial(wb,e7)).grid(row=0,column=3)
def wb(e7):
    cnt=len(sln)
    
    wb=Workbook()
    ws=wb.active
    file_name=fn.get()+'.xlsx'
    print(file_name)
    ws.title="Delivery Challan"
    ws.column_dimensions['A'].width =  11.57
    ws.column_dimensions['B'].width=19.29
    ws.column_dimensions['C'].width=8.29
    ws.column_dimensions['D'].width=35.57
    ws.column_dimensions['E'].width=6.29
    ws.column_dimensions['F'].width=5.71
    ws.column_dimensions['G'].width=12.14
    ws.column_dimensions['H'].width=30.86
    ws.merge_cells('A1:H1')
    ws.merge_cells('C13:D14')
    ws.merge_cells('A13:A14')
    ws.merge_cells('B13:B14')
    ws.merge_cells('E13:E14')
    ws.merge_cells('F13:F14')
    ws.merge_cells('G13:G14')
    ws.merge_cells('H13:H14')
    
    ws['A1']='Delivery Challan'
    ws.merge_cells('A2:H2')
    ws['A2']='Original/Duplicate/Triplicate'
    ws.merge_cells('B11:C11')
    ws['b11']='State:Karnataka'
    ws.merge_cells('E11:G11')
    ws['e11']='State:Karnataka'
    po=['PO No.:','PO Date:']
    
    pay="Payment/Delivery Terms:"
    payment=pay+e7.get()
    
    ws['d3']='Consignee:'
    t=0
    a=0
    while(t<len(con)):
        a=con[t].get()
        print(con[t].get())
        ws.cell(row=t+4, column=4).value=con[t].get()
        t+=1
    From=["Shipper / Exporters:","L&T Technology Services Limited - SEZ Unit II.,","“Hazel-Block L3”, Ground to 3rd Floors,","Manyata Embassy Business Park,","Nagawara Hobli, Outer Ring Road,","Bangalore - 560045","GST # 29AACCL4310P1Z9"]
    cl_H=["DC No.:LTTS/SEZ-II/024/2018","DC Date: 23.04.2019","Client Code:","Job Code:","Sales Order No.:","Place of Supply: Bangalore"]
    TnC="Other Terms & Conditions:The Above Items are sent for repeir & return purpose on Returnable basis and these items are not for Sale."
    bom=["S.N.","HSN Code","Goods/ Services Description","Qty","UOM","Unit Rate (INR)","Taxable Value(INR)"]
    sc="SC: 29"
    ws['A13']=bom[0]
    ws['B13']=bom[1]
    ws['C13']=bom[2]
    ws['E13']=bom[3]
    ws['F13']=bom[4]
    ws['G13']=bom[5]
    ws['h13']=bom[6]
    ws['a11']=ws['d11']=sc
    ws.merge_cells(start_row=15+cnt,start_column=1,end_row=15+cnt,end_column=7)
    ws.cell(row=15+cnt,column=1).value ='Total Amount of Taxable Value'
    ws.merge_cells(start_row=16+cnt,start_column=1,end_row=16+cnt,end_column=8)
    ws.merge_cells(start_row=17+cnt,start_column=1,end_row=17+cnt,end_column=8)
    ws.merge_cells(start_row=18+cnt,start_column=1,end_row=18+cnt,end_column=8)
    ws.cell(row=18+cnt,column=1).value="GSTIN: Applied for GST, Ref ARN:-AA290717006016A"
    
    ws.merge_cells(start_row=15+cnt+6,start_column=1,end_row=15+cnt+6+5,end_column=8)
    ws.cell(row=15+cnt+6,column=1).value=payment
    
    

    t=0
    while(t<len(From)):
        ws.cell(row=t+3, column=1).value=From[t]
        t+=1
        
    t=0
    while(t<len(cl_H)):
        
        ws.cell(row=t+3, column=8).value=cl_H[t]
        t+=1
        
    t=0
    while(t<len(po)):
        
        ws.cell(row=t+11, column=8).value=po[t]
        t+=1
    
        
    ws.merge_cells(start_row=15+cnt+6+5+1,start_column=1,end_row=15+cnt+6+5+1+7,end_column=5)
    ws.cell(row=15+cnt+6+5+1,column=1).value=TnC   
 
    
    
    ws.merge_cells(start_row=15+cnt+6+5+1,start_column=7,end_row=15+cnt+6+5+1+6,end_column=8)
    ws.cell(row=15+cnt+6+5+1,column=7).value="For L&T Technology Services Ltd - SEZ Unit II.,"
    
    ws.merge_cells(start_row=15+cnt+6+5+1+7,start_column=7,end_row=15+cnt+6+5+1+7,end_column=8)
    ws.cell(row=15+cnt+6+5+1+7,column=7).value="Authorized Signatory"
    i=0
    while ((i+1)<=cnt):
        a=sln[i]
        b=hsn[i].get()
        cd=desc[i].get()
        e=qty[i].get()
        f=uom[i].get()
        g=rate[i].get()
        ws.cell(row=14+i+1,column=1).value=a #sln[i].get()
        ws.cell(row=14+i+1,column=2).value=b #hsn[i].get()
        ws.merge_cells(start_row=14+i+1,start_column=3,end_row=14+i+1,end_column=4)
        ws.cell(row=14+i+1,column=3).value=cd #desc[i].get()
        ws.cell(row=14+i+1,column=5).value=e #qty[i].get()
        ws.cell(row=14+i+1,column=6).value=f #uom[i].get()
        ws.cell(row=14+i+1,column=7).value=g #rate[i].get()
        k=14+i+1
        ws.cell(row=14+i+1,column=8).value="=E"+str(k)+"*"+"G"+str(k)      
        i+=1
    y=14+cnt
    ws.cell(row=15+cnt,column=8).value ="=SUM(H15"+":"+"H"+str(y)+")"
    wb.save(file_name)
    
    


tk.Label(root,text="Number of Parts").grid(row=0,column=0)
e1=tk.Entry(root)
e1.grid(row=0,column=1)
tk.Label(root,text="File Name").grid(row=1,column=0)
fn=tk.Entry(root)
fn.grid(row=1,column=1)
tk.Button(root,text='Add',command=partial(add,e1)).grid(row=0,column=2)
tk.Button(root,text="Close",command=root.destroy).grid(row=0, column=4)
root.mainloop()
