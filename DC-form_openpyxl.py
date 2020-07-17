import tkinter as tk
import xlsxwriter
import inflect
from openpyxl import Workbook
from functools import partial

root=tk.Tk()
root.title("DC")
sln=[]
hsn=[]
desc=[]
qty=[]
uom=[]
rate=[]
con=[]

def add(e1):
    n=e1.get()
    x=int(n)
    print(x)
    i=1
    #master=tk.Toplevel(root)
    tk.Label(root,text="Payment/Delivery Terms:").grid(row=2,column=0)
    e7=tk.Entry(root)
    e7.grid(row=2,column=1)
    tk.Label(root,text="Consignee:").grid(row=1,column=3)
    p=1
    while (p<=7):
        ent=tk.Entry(root)
        ent.grid(row=p+1,column=3)
        con.append(ent)
        p+=1
    
    tk.Label(root,text="S.N.").grid(row=9,column=0)
    tk.Label(root,text="HSN Code").grid(row=9,column=1)
    tk.Label(root,text="Goods/ Services Description").grid(row=9,column=2)
    tk.Label(root,text="Qty").grid(row=9,column=3)
    tk.Label(root,text="UOM").grid(row=9,column=4)
    tk.Label(root,text="Unit Rate (INR)").grid(row=9,column=5)
    
    while(i<=x):
            sln.append(i)
            tk.Label(root,text=i).grid(row=i+9,column=0)
            e2=tk.Entry(root)
            e2.grid(row=i+9,column=1)
            hsn.append(e2)
            e3=tk.Entry(root)
            e3.grid(row=i+9,column=2)
            desc.append(e3)
            e4=tk.Entry(root)
            e4.grid(row=i+9,column=3)
            qty.append(e4)
            e5=tk.Entry(root)
            e5.grid(row=i+9,column=4)
            uom.append(e5)
            e6=tk.Entry(root)
            e6.grid(row=i+9,column=5)
            rate.append(e6)
            i+=1
    tk.Button(root,text='Next',command=partial(wb,e7)).grid(row=0,column=3)
def wb(e7):
    cnt=len(sln)
    
    wb=Workbook()
    ws=wb.active
    file_name=fn.get()+'.xlsx'
    print(file_name)
    ws.title="Delivery Challan"
    ws.column_dimensions['A'].width =  11.57
    ws.column_dimensions['B'].width=19.29
    ws.column_dimensions['C'].width=8.29
    ws.column_dimensions['D'].width=35.57
    ws.column_dimensions['E'].width=6.29
    ws.column_dimensions['F'].width=5.71
    ws.column_dimensions['G'].width=12.14
    ws.column_dimensions['H'].width=30.86
    ws.merge_cells('A1:H1')
    ws.merge_cells('C13:D14')
    ws.merge_cells('A13:A14')
    ws.merge_cells('B13:B14')
    ws.merge_cells('E13:E14')
    ws.merge_cells('F13:F14')
    ws.merge_cells('G13:G14')
    ws.merge_cells('H13:H14')
    
    ws['A1']='Delivery Challan'
    ws.merge_cells('A2:H2')
    ws['A2']='Original/Duplicate/Triplicate'
    ws.merge_cells('B11:C11')
    ws['b11']='State:Karnataka'
    ws.merge_cells('E11:G11')
    ws['e11']='State:Karnataka'
    po=['PO No.:','PO Date:']
    
    pay="Payment/Delivery Terms:"
    payment=pay+e7.get()
    
    ws['d3']='Consignee:'
    t=0
    a=0
    while(t<len(con)):
        a=con[t].get()
        print(con[t].get())
        ws.cell(row=t+4, column=4).value=con[t].get()
        t+=1
    From=["Shipper / Exporters:","L&T Technology Services Limited - SEZ Unit II.,","“Hazel-Block L3”, Ground to 3rd Floors,","Manyata Embassy Business Park,","Nagawara Hobli, Outer Ring Road,","Bangalore - 560045","GST # 29AACCL4310P1Z9"]
    cl_H=["DC No.:LTTS/SEZ-II/024/2018","DC Date: 23.04.2019","Client Code:","Job Code:","Sales Order No.:","Place of Supply: Bangalore"]
    TnC="Other Terms & Conditions:The Above Items are sent for repeir & return purpose on Returnable basis and these items are not for Sale."
    bom=["S.N.","HSN Code","Goods/ Services Description","Qty","UOM","Unit Rate (INR)","Taxable Value(INR)"]
    ws['A13']=bom[0]
    ws['B13']=bom[1]
    ws['C13']=bom[2]
    ws['E13']=bom[3]
    ws['F13']=bom[4]
    ws['G13']=bom[5]
    ws['h13']=bom[6]
    ws['A15']='Total Amount of Taxable Value'
    ws.insert_rows(15,amount=cnt)
    ws.merge_cells(start_row=21+cnt,start_column=1,end_row=21+5+cnt,end_column=8)
    ws.cell(row=21+cnt,column=1).value=payment
    

    
    

    
    print(len(From))
    t=0
    while(t<len(From)):
        ws.cell(row=t+3, column=1).value=From[t]
        t+=1
        
    t=0
    while(t<len(cl_H)):
        
        ws.cell(row=t+3, column=8).value=cl_H[t]
        t+=1
        
    t=0
    while(t<len(po)):
        
        ws.cell(row=t+11, column=8).value=po[t]
        t+=1
    wb.save(file_name)
    ws.cell(row=26+cnt,column=7).value=0
    ws.merge_cells(start_row=26+cnt,start_column=7,end_row=26+6+cnt,end_column=8)
    ws.cell(row=26+cnt,column=7).value="For L&T Technology Services Ltd - SEZ Unit II.,"
    ws.merge_cells(start_row=26+cnt,start_column=1,end_row=26+7+cnt,end_column=5)
    ws.cell(row=26+cnt,column=1).value=TnC   
    
    wb.save(file_name)
    
    


tk.Label(root,text="Number of Parts").grid(row=0,column=0)
e1=tk.Entry(root)
e1.grid(row=0,column=1)
tk.Label(root,text="File Name").grid(row=1,column=0)
fn=tk.Entry(root)
fn.grid(row=1,column=1)
tk.Button(root,text='Add',command=partial(add,e1)).grid(row=0,column=2)
tk.Button(root,text="Close",command=root.destroy).grid(row=0, column=4)
root.mainloop()
